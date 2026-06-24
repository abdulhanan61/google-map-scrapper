"""
Eye Clinic Scraper - District Gujrat, Punjab, Pakistan
Comprehensive search across all cities, towns, and keywords
Uses Selenium with your existing Chrome browser
Saves results to Excel (.xlsx)

Requirements:
    pip install selenium openpyxl webdriver-manager
"""

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# All cities/towns in District Gujrat x multiple keywords
LOCATIONS = [
    "Gujrat City", "Kharian", "Lalamusa", "Sarai Alamgir",
    "Jalalpur Jattan", "Kunjah", "Dinga", "Waziabad",
    "Bharowal", "Kharial", "Mangla", "Lala Musa",
    "Phalia", "Kot Isa Khan", "Sukheke", "Bhimber Road Gujrat"
]

KEYWORDS = [
    "eye clinic",
    "eye hospital",
    "eye care",
    "ophthalmic clinic",
    "chasma centre",
    "chasma ghar",
]

OUTPUT_FILE = "gujrat_district_eye_clinics.xlsx"


def build_queries():
    queries = []
    for location in LOCATIONS:
        for keyword in KEYWORDS:
            queries.append(f"{keyword} {location} Punjab Pakistan")
    return queries


def get_driver():
    options = Options()
    # options.add_argument("--headless")  # Uncomment to run silently
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    return driver


def scroll_results(driver):
    try:
        panel = driver.find_element(By.CSS_SELECTOR, 'div[role="feed"]')
        last_count = 0
        no_change = 0
        while no_change < 5:
            driver.execute_script("arguments[0].scrollBy(0, 1500);", panel)
            time.sleep(2)
            items = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
            current_count = len(items)
            if current_count == last_count:
                no_change += 1
            else:
                no_change = 0
                last_count = current_count
    except Exception as e:
        print(f"  Scroll error: {e}")


def get_text_safe(driver, selectors):
    for selector in selectors:
        try:
            el = driver.find_element(By.CSS_SELECTOR, selector)
            text = el.text.strip()
            if text:
                return text
        except:
            continue
    return "N/A"


def scrape_query(driver, query, seen_names):
    results = []
    url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
    print(f"\n🔍 {query}")
    driver.get(url)
    time.sleep(4)

    scroll_results(driver)

    link_elements = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
    hrefs = list(dict.fromkeys([
        el.get_attribute("href") for el in link_elements
        if el.get_attribute("href")
    ]))

    print(f"  → {len(hrefs)} listings found")

    for i, href in enumerate(hrefs):
        try:
            driver.get(href)
            time.sleep(2.5)

            name = get_text_safe(driver, ['h1', 'h1.DUwDvf'])
            address = get_text_safe(driver, [
                'button[data-item-id="address"]',
                '[data-tooltip="Copy address"]',
                'button[aria-label*="Address"]',
                '.Io6YTe',
            ])

            if name and name != "N/A":
                key = name.lower().strip()
                if key not in seen_names:
                    seen_names.add(key)
                    results.append({"name": name, "address": address})
                    print(f"  ✅ [{i+1}] {name} — {address}")
                else:
                    print(f"  ⏭️  [{i+1}] Duplicate skipped: {name}")
            else:
                print(f"  ❌ [{i+1}] Skipped (no name)")

        except Exception as e:
            print(f"  ⚠️  [{i+1}] Error: {e}")
            continue

    return results


def scrape_all():
    all_clinics = []
    seen_names = set()
    queries = build_queries()
    driver = get_driver()

    print(f"Total searches to run: {len(queries)}")
    print("=" * 60)

    try:
        for idx, query in enumerate(queries, 1):
            print(f"\n[{idx}/{len(queries)}]", end=" ")
            results = scrape_query(driver, query, seen_names)
            all_clinics.extend(results)
            print(f"  Running total: {len(all_clinics)} unique clinics")

    finally:
        driver.quit()

    return all_clinics


def save_to_excel(clinics, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Eye Clinics - District Gujrat"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["#", "Clinic Name", "Address"]
    widths = [5, 45, 65]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    for i, clinic in enumerate(clinics, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=clinic["name"]).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=2).alignment = left
        ws.cell(row=row, column=3, value=clinic["address"]).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=3).alignment = left

        if i % 2 == 0:
            fill = PatternFill("solid", start_color="D6E4F0")
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = fill

    total_row = len(clinics) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=2, value=f"=COUNTA(B2:B{len(clinics)+1})").font = Font(name="Arial", bold=True)

    wb.save(output_file)
    print(f"\n✅ Saved to: {output_file}")
    print(f"   Total unique clinics: {len(clinics)}")


if __name__ == "__main__":
    print("=" * 60)
    print("Eye Clinic Scraper — District Gujrat, Pakistan")
    print(f"Cities: {len(LOCATIONS)} | Keywords: {len(KEYWORDS)}")
    print("=" * 60)

    clinics = scrape_all()

    if clinics:
        save_to_excel(clinics, OUTPUT_FILE)
    else:
        print("No clinics found. Check your internet connection and try again.")
