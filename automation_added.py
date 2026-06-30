import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
import time
from plyer import notification

EXCEL_FILE = r"C:\Users\UseR\Desktop\coding\eye clinic\gujrat_eye_clinics_with_phones.xlsx"
SEARCH_QUERY = [
    "eye clinics in Gujrat Punjab Pakistan",
    "eye clinics in Kharian Punjab Pakistan",
    "eye clinics in Jalalpur Jattan Pakistan",
    "eye clinics in Lalamusa Pakistan",
    "eye clinics in Sarai Alamgir, Pakistnan",
    "eye clinics in Kunjah pakistan",
    "eye clinic in Dinga Pakistan",
    "eye clinic in Waziabad Pakistan",
]

YELLOW_FILL = PatternFill("solid", fgColor="FFD700")
NEW_FONT    = Font(bold=True, name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def scrape_google_maps(query):
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    # REMOVED --headless so Google doesn't block it
    
    driver = webdriver.Chrome(options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.get(f"https://www.google.com/maps/search/{query.replace(' ', '+')}")
    time.sleep(5)

    results = []
    try:
        scrollable = driver.find_element(By.XPATH, '//div[@role="feed"]')
        for _ in range(10):
            driver.execute_script("arguments[0].scrollTop += 1000", scrollable)
            time.sleep(1.5)
    except:
        print("⚠️ Could not find listings feed — page may not have loaded correctly.")
        driver.quit()
        return pd.DataFrame(columns=["Clinic Name", "Phone Number", "Address"])

    listings = driver.find_elements(By.CLASS_NAME, "hfpxzc")
    print(f"Found {len(listings)} listings on Google Maps")

    for listing in listings:
        try:
            listing.click()
            time.sleep(2)
            name = driver.find_element(By.CLASS_NAME, "DUwDvf").text
            try:
                phone = driver.find_element(
                    By.XPATH, '//button[contains(@data-item-id,"phone")]'
                ).get_attribute("data-item-id").replace("phone:tel:", "")
            except:
                phone = "N/A"
            try:
                address = driver.find_element(
                    By.XPATH, '//button[@data-item-id="address"]').text
            except:
                address = "N/A"
            results.append({"Clinic Name": name, "Phone Number": phone, "Address": address})
            print(f"  ✓ {name}")
        except:
            continue

    driver.quit()
    
    if not results:
        print("⚠️ Scraper got 0 results — Google may have changed its page structure.")
        return pd.DataFrame(columns=["Clinic Name", "Phone Number", "Address"])
    
    return pd.DataFrame(results)


def add_new_to_excel(new_df):
    print(f"\nScraped {len(new_df)} total listings")
    
    if new_df.empty:
        print("⚠️ No data scraped. Exiting.")
        return

    # Read existing — specify sheet name from your file
    existing = pd.read_excel(EXCEL_FILE, sheet_name=0)
    print("Existing columns:", existing.columns.tolist())
    
    # Get the clinic name column (handle whatever it's called)
    name_col = [c for c in existing.columns if "name" in c.lower() or "clinic" in c.lower()]
    if not name_col:
        print("❌ Could not find Clinic Name column in Excel. Columns:", existing.columns.tolist())
        return
    name_col = name_col[0]
    print(f"Using column: '{name_col}'")

    existing_names = existing[name_col].astype(str).str.lower().str.strip().tolist()
    new_only = new_df[~new_df["Clinic Name"].str.lower().str.strip().isin(existing_names)]

    if new_only.empty:
        print("✅ No new clinics found — dataset is up to date.")
        return

    print(f"🆕 {len(new_only)} new clinics found!")

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    today = datetime.now().strftime("%Y-%m-%d")
    last_row = ws.max_row

    for i, (_, row) in enumerate(new_only.iterrows()):
        r = last_row + i + 1
        values = [r - 1, f" NEW — {row['Clinic Name']}", row['Address'], row['Phone Number'], today]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = YELLOW_FILL
            cell.font      = NEW_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb.save(EXCEL_FILE)    # ← 4 spaces
    print(f"✅ Done! {len(new_only)} new clinics added and highlighted in yellow.")
    notification.notify(
        title="🏥 New Eye Clinics Found!",
        message=f"{len(new_only)} new clinics added to Excel.",
        timeout=10
    )                      # ← 4 spaces


# --- RUN ---
all_results = []
for query in SEARCH_QUERY:
    print(f"\n🔍 Searching: {query}")
    scraped = scrape_google_maps(query)
    all_results.append(scraped)

combined = pd.concat(all_results, ignore_index=True)
add_new_to_excel(combined)
# automatically run every week
# import schedule

# def run_scraper():
#     all_results = []
#     for query in SEARCH_QUERY:
#         print(f"\n🔍 Searching: {query}")
#         scraped = scrape_google_maps(query)
#         all_results.append(scraped)
    
#     combined = pd.concat(all_results, ignore_index=True)
#     add_new_to_excel(combined)

# # Run once immediately, then every week
# run_scraper()
# schedule.every().week.do(run_scraper)

# while True:
#     schedule.run_pending()
#     time.sleep(60)