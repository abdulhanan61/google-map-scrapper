"""
Phone Number Scraper for Eye Clinics - District Gujrat
Reads the cleaned Excel file, searches each clinic on Google Maps,
extracts phone numbers, and saves updated Excel with Phone Number column.

"""

import time
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

INPUT_FILE  = "gujrat_district_eye_clinics_cleaned.xlsx"
OUTPUT_FILE = "gujrat_eye_clinics_with_phones.xlsx"


def get_driver():
    options = Options()
    # options.add_argument("--headless")  # uncomment to run silently
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Chrome(service=Service("chromedriver.exe"), options=options)


def get_text_safe(driver, selectors):
    for sel in selectors:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            text = el.text.strip()
            if text:
                return text
        except:
            continue
    return None


def extract_phone(driver):
    # Try direct phone button selectors
    phone = get_text_safe(driver, [
        'button[data-item-id^="phone"]',
        'button[aria-label^="Phone"]',
        '[data-tooltip="Copy phone number"]',
        'a[href^="tel:"]',
    ])
    if phone:
        # Clean up: keep only digits, +, -, spaces, ()
        phone = re.sub(r'[^\d\+\-\(\) ]', '', phone).strip()
        if len(phone) >= 7:
            return phone

    # Fallback: scan page source for Pakistani phone patterns
    try:
        page = driver.page_source
        patterns = [
            r'(?:0|\+92)3\d{2}[\s\-]?\d{7}',   # mobile: 03xx-xxxxxxx
            r'(?:0|\+92)5\d{1,2}[\s\-]?\d{6,7}', # landline: 053-xxxxxxx
            r'\+92\s?\d{2,3}[\s\-]?\d{6,7}',
        ]
        for pat in patterns:
            match = re.search(pat, page)
            if match:
                return match.group().strip()
    except:
        pass

    return "N/A"


def search_clinic_on_maps(driver, name, address):
    query = f"{name} {address} Pakistan"
    url = "https://www.google.com/maps/search/" + query.replace(" ", "+")
    driver.get(url)
    time.sleep(3)

    # If multiple results appear, click the first one
    try:
        results = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
        if results:
            results[0].click()
            time.sleep(3)
    except:
        pass

    return extract_phone(driver)


def save_excel(df, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Eye Clinics - District Gujrat"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    green_fill   = PatternFill("solid", start_color="E2EFDA")  # has phone
    alt_fill     = PatternFill("solid", start_color="D6E4F0")
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    headers = ["#", "Clinic Name", "Address", "Phone Number"]
    widths  = [5, 42, 60, 20]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for i, row in df.iterrows():
        r = i + 2
        has_phone = str(row.get('Phone Number', 'N/A')).strip() not in ('N/A', '', 'nan')

        ws.cell(row=r, column=1, value=row['#']).alignment = center
        ws.cell(row=r, column=2, value=row['Clinic Name']).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2).alignment = left
        ws.cell(row=r, column=3, value=row['Address']).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=3).alignment = left
        ws.cell(row=r, column=4, value=row.get('Phone Number', 'N/A')).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=4).alignment = center

        # Green row if phone found, alternating blue otherwise
        fill = green_fill if has_phone else (alt_fill if (i+1) % 2 == 0 else None)
        if fill:
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = fill

    # Summary
    total_row = len(df) + 2
    found = df[df['Phone Number'] != 'N/A'].shape[0] if 'Phone Number' in df.columns else 0
    ws.cell(row=total_row, column=1, value="Total").font   = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=2, value=f"=COUNTA(B2:B{len(df)+1})").font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=3, value=f"Phones found: {found}").font    = Font(name="Arial", bold=True, color="375623")

    wb.save(output_file)


def main():
    print("=" * 60)
    print("Phone Number Scraper — District Gujrat Eye Clinics")
    print("=" * 60)

    df = pd.read_excel(INPUT_FILE)
    df = df[df['Clinic Name'].notna() & (df['Clinic Name'].astype(str) != 'Total')].reset_index(drop=True)
    df['Phone Number'] = 'N/A'

    driver = get_driver()
    found_count = 0

    try:
        for i, row in df.iterrows():
            name    = str(row['Clinic Name'])
            address = str(row['Address'])

            print(f"\n[{i+1}/{len(df)}] {name}")
            phone = search_clinic_on_maps(driver, name, address)
            df.at[i, 'Phone Number'] = phone

            if phone != 'N/A':
                found_count += 1
                print(f"  ✅ {phone}")
            else:
                print(f"  ❌ Not found")

            # Save progress every 10 clinics
            if (i + 1) % 10 == 0:
                save_excel(df, OUTPUT_FILE)
                print(f"\n  💾 Progress saved ({i+1}/{len(df)} done, {found_count} phones found)\n")

    finally:
        driver.quit()
        save_excel(df, OUTPUT_FILE)

    print("\n" + "=" * 60)
    print(f"✅ Done! {found_count}/{len(df)} phone numbers found")
    print(f"📁 Saved to: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
